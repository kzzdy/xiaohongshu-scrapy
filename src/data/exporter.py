"""
数据导出器模块

支持将数据导出为多种格式：Excel、JSON、CSV
"""

import json
import csv
from typing import List, Dict, Any, Optional
from pathlib import Path
from enum import Enum
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from loguru import logger

from .validator import DataValidator


class ExportFormat(Enum):
    """导出格式枚举"""

    EXCEL = "excel"
    JSON = "json"
    CSV = "csv"


class DataExporter:
    """数据导出器"""

    # 笔记数据的列标题
    NOTE_HEADERS = [
        "笔记id",
        "笔记url",
        "笔记类型",
        "用户id",
        "用户主页url",
        "昵称",
        "头像url",
        "标题",
        "描述",
        "点赞数量",
        "收藏数量",
        "评论数量",
        "分享数量",
        "视频封面url",
        "视频地址url",
        "图片地址url列表",
        "标签",
        "上传时间",
        "ip归属地",
    ]

    # 用户数据的列标题
    USER_HEADERS = [
        "用户id",
        "用户主页url",
        "用户名",
        "头像url",
        "小红书号",
        "性别",
        "ip地址",
        "介绍",
        "关注数量",
        "粉丝数量",
        "作品被赞和收藏数量",
        "标签",
    ]

    # 评论数据的列标题
    COMMENT_HEADERS = [
        "笔记id",
        "笔记url",
        "评论id",
        "用户id",
        "用户主页url",
        "昵称",
        "头像url",
        "评论内容",
        "评论标签",
        "点赞数量",
        "上传时间",
        "ip归属地",
        "图片地址url列表",
    ]

    def __init__(self, output_dir: str = "datas/excel_datas"):
        """
        初始化数据导出器

        Args:
            output_dir: 输出目录路径
        """
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)
        self.validator = DataValidator()

    def export(
        self,
        data: List[Dict[str, Any]],
        filename: str,
        format: ExportFormat = ExportFormat.EXCEL,
        data_type: str = "note",
    ) -> str:
        """
        导出数据

        Args:
            data: 要导出的数据列表
            filename: 文件名（不含扩展名）
            format: 导出格式
            data_type: 数据类型 ('note', 'user', 'comment')

        Returns:
            导出文件的完整路径
        """
        if not data:
            logger.warning("No data to export")
            return ""

        # 清理文件名
        filename = self.validator.clean_filename(filename)

        # 根据格式选择输出目录和导出方法
        # 确保使用正确的基础目录（如果output_dir是datas，则直接使用；如果是datas/excel_datas，则使用parent）
        base_dir = self.output_dir
        if base_dir.name in ["excel_datas", "json_datas", "csv_datas"]:
            base_dir = base_dir.parent
        
        if format == ExportFormat.EXCEL:
            output_dir = base_dir / "excel_datas"
            output_dir.mkdir(parents=True, exist_ok=True)
            filepath = output_dir / f"{filename}.xlsx"
            self.export_to_excel(data, str(filepath), data_type)
        elif format == ExportFormat.JSON:
            output_dir = base_dir / "json_datas"
            output_dir.mkdir(parents=True, exist_ok=True)
            filepath = output_dir / f"{filename}.json"
            self.export_to_json(data, str(filepath))
        elif format == ExportFormat.CSV:
            output_dir = base_dir / "csv_datas"
            output_dir.mkdir(parents=True, exist_ok=True)
            filepath = output_dir / f"{filename}.csv"
            self.export_to_csv(data, str(filepath), data_type)
        else:
            raise ValueError(f"Unsupported export format: {format}")

        # 显示导出统计信息
        stats = self.get_export_stats(str(filepath), len(data))
        logger.info(f"数据导出成功: {stats['filepath']}")
        logger.info(f"导出记录数: {stats['record_count']}, 文件大小: {stats['file_size']}")

        return str(filepath)

    def export_to_excel(
        self, data: List[Dict[str, Any]], filepath: str, data_type: str = "note"
    ) -> None:
        """
        导出数据为Excel格式

        Args:
            data: 要导出的数据列表
            filepath: 输出文件路径
            data_type: 数据类型 ('note', 'user', 'comment')
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "数据"

        # 根据数据类型选择表头
        if data_type == "note":
            headers = self.NOTE_HEADERS
        elif data_type == "user":
            headers = self.USER_HEADERS
        elif data_type == "comment":
            headers = self.COMMENT_HEADERS
        else:
            # 如果类型未知，使用数据的键作为表头
            headers = list(data[0].keys()) if data else []

        # 写入表头并设置样式
        ws.append(headers)
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        header_alignment = Alignment(horizontal="center", vertical="center")

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        # 写入数据行
        for row_data in data:
            # 清理数据中的非法字符
            cleaned_data = {
                k: self.validator.clean_text_for_excel(str(v)) for k, v in row_data.items()
            }
            ws.append(list(cleaned_data.values()))

        # 自动调整列宽
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)  # 最大宽度50
            ws.column_dimensions[column_letter].width = adjusted_width

        # 保存文件
        wb.save(filepath)
        logger.debug(f"Excel file saved: {filepath}")

    def export_to_json(self, data: List[Dict[str, Any]], filepath: str, indent: int = 2) -> None:
        """
        导出数据为JSON格式

        Args:
            data: 要导出的数据列表
            filepath: 输出文件路径
            indent: JSON缩进空格数
        """
        with open(filepath, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=indent)
        logger.debug(f"JSON file saved: {filepath}")

    def export_to_csv(
        self, data: List[Dict[str, Any]], filepath: str, data_type: str = "note"
    ) -> None:
        """
        导出数据为CSV格式

        Args:
            data: 要导出的数据列表
            filepath: 输出文件路径
            data_type: 数据类型 ('note', 'user', 'comment')
        """
        if not data:
            return

        # 根据数据类型选择表头
        if data_type == "note":
            headers = self.NOTE_HEADERS
            fieldnames = list(data[0].keys())
        elif data_type == "user":
            headers = self.USER_HEADERS
            fieldnames = list(data[0].keys())
        elif data_type == "comment":
            headers = self.COMMENT_HEADERS
            fieldnames = list(data[0].keys())
        else:
            # 如果类型未知，使用数据的键
            fieldnames = list(data[0].keys())
            headers = fieldnames

        with open(filepath, "w", encoding="utf-8-sig", newline="") as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames)

            # 写入表头
            writer.writerow(dict(zip(fieldnames, headers)))

            # 写入数据
            for row_data in data:
                # 清理数据
                cleaned_data = {
                    k: self.validator.clean_text_for_excel(str(v)) for k, v in row_data.items()
                }
                writer.writerow(cleaned_data)

        logger.debug(f"CSV file saved: {filepath}")

    def get_export_stats(self, filepath: str, record_count: int) -> Dict[str, Any]:
        """
        获取导出统计信息

        Args:
            filepath: 文件路径
            record_count: 记录数量

        Returns:
            包含统计信息的字典
        """
        file_path = Path(filepath)

        if not file_path.exists():
            return {
                "filepath": filepath,
                "record_count": record_count,
                "file_size": "N/A",
                "file_size_bytes": 0,
            }

        file_size_bytes = file_path.stat().st_size

        # 格式化文件大小
        if file_size_bytes < 1024:
            file_size = f"{file_size_bytes} B"
        elif file_size_bytes < 1024 * 1024:
            file_size = f"{file_size_bytes / 1024:.2f} KB"
        else:
            file_size = f"{file_size_bytes / (1024 * 1024):.2f} MB"

        return {
            "filepath": filepath,
            "record_count": record_count,
            "file_size": file_size,
            "file_size_bytes": file_size_bytes,
            "export_time": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        }

    def export_notes(
        self,
        notes: List[Dict[str, Any]],
        filename: Optional[str] = None,
        format: ExportFormat = ExportFormat.EXCEL,
    ) -> str:
        """
        导出笔记数据的便捷方法

        Args:
            notes: 笔记数据列表
            filename: 文件名（可选，默认使用时间戳）
            format: 导出格式

        Returns:
            导出文件路径
        """
        if filename is None:
            filename = f"notes_{datetime.now().strftime('%Y%m%d_%H%M%S')}"

        return self.export(notes, filename, format, data_type="note")

    def export_users(
        self,
        users: List[Dict[str, Any]],
        filename: Optional[str] = None,
        format: ExportFormat = ExportFormat.EXCEL,
    ) -> str:
        """
        导出用户数据的便捷方法

        Args:
            users: 用户数据列表
            filename: 文件名（可选，默认使用时间戳）
            format: 导出格式

        Returns:
            导出文件路径
        """
        if filename is None:
            filename = f"users_{datetime.now().strftime('%Y%m%d_%H%M%S')}"

        return self.export(users, filename, format, data_type="user")

    def export_comments(
        self,
        comments: List[Dict[str, Any]],
        filename: Optional[str] = None,
        format: ExportFormat = ExportFormat.EXCEL,
    ) -> str:
        """
        导出评论数据的便捷方法

        Args:
            comments: 评论数据列表
            filename: 文件名（可选，默认使用时间戳）
            format: 导出格式

        Returns:
            导出文件路径
        """
        if filename is None:
            filename = f"comments_{datetime.now().strftime('%Y%m%d_%H%M%S')}"

        return self.export(comments, filename, format, data_type="comment")
